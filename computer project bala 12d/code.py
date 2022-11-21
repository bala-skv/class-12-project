#project to create a exam window
from tkinter import ttk
from tkinter import *
from PIL import Image,ImageTk
import random
import openpyxl 
from openpyxl import workbook,load_workbook
wp=load_workbook('physics.xlsx')
wps=wp.active
wm=load_workbook('maths.xlsx')
wms=wm.active
wc=load_workbook('chemistry.xlsx')
wcs=wc.active
root=Tk()
'''ws['F3'].value='test'
q=int(input('enter integer'))
print(ws['{1}{0}'.format(q,'B')].value)
wb.save('physics.xlsx')'''        
    def image_display():
        image_path=r"C:\Users\balas\Desktop\computer project\QUESTION IMAGES\{0}.png".format(q)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.grid(column=0,row=0)
    def mcq():
        image_path=r"C:\Users\balas\Desktop\computer project\QUESTION IMAGES\{0}.png".format(q)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.grid(column=0,row=0)
        radio=IntVar()
        def selection():
            markslstmcq.append([c,radio.get(),a])
        v = StringVar(root, "1")
        values = [["A","A"],["B","B"],["C","C"],["D","D"]]
        iteration_no=0
        for (text, value) in values:
            iteration_no+=1
            globals()[f'Radiobutton{iteration_no}']=Radiobutton(root, text = text, variable = v,
                    value = value,command=selection)
            globals()[f'Radiobutton{iteration_no}'].grid(row=iteration_no,column=3)
        def deletion():
            Radiobutton1.destroy()
            Radiobutton2.destroy()
            Radiobutton3.destroy()
            Radiobutton4.destroy()
            label.destroy()
            
        selection()   
        button1=Button(root,text='next',command=deletion).grid(row=10,column=3)
        
    def integer():
        image_path=r"C:\Users\balas\Desktop\computer project\QUESTION IMAGES\{0}.png".format(q)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.grid(column=0,row=0)
        def submit():
            answer=entryvariable.get()
            markslstint.append([c,answer,a])
        entryvariable=StringVar(root)
        answer_entry = Entry(root,textvariable = entryvariable)
        answer_entry.grid(row=2,column=2)
        enter_button=Button(root,text = 'ENTER', command = submit)
        enter_button.grid(row=2,column=3)
        def deletion():
            label.destroy()
            enter_button.destroy()
        button1=Button(root,text='next',command=deletion).grid(row=10,column=3)
        submit()
            
        
    lst=[]
    markslstmcq=[]
    markslstint=[]
    for i in range(6):
            r=random.randrange(2,31)
            while r not in lst:
                lst.append(r)
                q=wps['A{0}'.format(r)].value
                c=wps['B{0}'.format(r)].value
                a=wps['C{0}'.format(r)].value
                t=wps['D{0}'.format(r)].value
                if t=='m':
                                    
                elif t=='i':
                    
quatertest()
                
                    

                    
                
                    
                    
                    
                
                
                
            
    
    
          

 
