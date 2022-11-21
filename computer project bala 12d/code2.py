from tkinter import *
from tkinter import ttk
from PIL import Image,ImageTk
import openpyxl 
from openpyxl import workbook,load_workbook
import random
wp=load_workbook(r'C:\Users\Dell\Desktop\computer project bala 12d\physics.xlsx')
wps=wp.active
wm=load_workbook(r'C:\Users\Dell\Desktop\computer project bala 12d\maths.xlsx')
wms=wm.active
wc=load_workbook(r'C:\Users\Dell\Desktop\computer project bala 12d\chemistry.xlsx')
wcs=wc.active

count=0
lst=[]
phymarkslstmcq=[]
phymarkslstint=[]
chemmarkslstmcq=[]
mathsmarkslstmcq=[]
chemmarkslstint=[]
mathsmarkslstint=[]
for i in range(3):
    r=random.randrange(2,31)
    while r not in lst:
        lst.append(r)
        q=wps['A{0}'.format(r)].value
        c=wps['B{0}'.format(r)].value
        a=wps['C{0}'.format(r)].value
        t=wps['D{0}'.format(r)].value
        root=Tk()
        root.geometry("1000x1000")
        print(q)
        image_path=r"C:\Users\Dell\Desktop\computer project bala 12d\QUESTION IMAGES\{0}.png".format(q)
        print(image_path)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.pack(anchor=N)
        if t=='m':
            radio=StringVar()
            radio.set('A')
            values = [["A","A"],["B","B"],["C","C"],["D","D"]]
            def clicked(option):
                phymarkslstmcq.append([c,option,a,q])
            iteration_no=0
            for (text, val) in values:
                iteration_no+=1
                Radiobutton(root, text = text, variable = radio,value = val).pack(anchor=S)
            Button(root,text='save option',command=lambda: clicked(radio.get())).pack(anchor=S)
        elif t=='i':
            def submit():
                answer=entryvariable.get()
                phymarkslstint.append([c,answer,a,q])
            entryvariable=StringVar(root)
            answer_entry = Entry(root,textvariable = entryvariable)
            answer_entry.pack(anchor=S)
            enter_button=Button(root,text = 'ENTER', command = submit)
            enter_button.pack(anchor=S)
        button1=Button(root,text='next',command=root.destroy).pack(anchor=S)
        root.mainloop()
lst=[]
for i in range(3):
    r=random.randrange(2,31)
    while r not in lst:
        lst.append(r)
        q=wcs['A{0}'.format(r)].value
        c=wcs['B{0}'.format(r)].value
        a=wcs['C{0}'.format(r)].value
        t=wcs['D{0}'.format(r)].value
        root=Tk()
        root.geometry("1000x1000")
        print(q)
        image_path=r"C:\Users\Dell\Desktop\computer project bala 12d\QUESTION IMAGES\{0}.png".format(q)
        print(image_path)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.pack(anchor=N)
        if t=='m':
            radio=StringVar()
            radio.set('A')
            values = [["A","A"],["B","B"],["C","C"],["D","D"]]
            def clicked(option):
                chemmarkslstmcq.append([c,option,a,q])
            iteration_no=0
            for (text, val) in values:
                iteration_no+=1
                Radiobutton(root, text = text, variable =radio,value = val).pack(anchor=S)
            Button(root,text='save option',command=lambda: clicked(radio.get())).pack(anchor=S)
        elif t=='i':
            def submit():
                answer=entryvariable.get()
                chemmarkslstint.append([c,answer,a,q])
            entryvariable=StringVar(root)
            answer_entry = Entry(root,textvariable = entryvariable)
            answer_entry.pack(anchor=S)
            enter_button=Button(root,text = 'ENTER', command = submit)
            enter_button.pack(anchor=S)
        button1=Button(root,text='next',command=root.destroy).pack(anchor=S)
        root.mainloop()
lst=[]
for i in range(3):
    r=random.randrange(2,31)
    while r not in lst:
        lst.append(r)
        q=wms['A{0}'.format(r)].value
        c=wms['B{0}'.format(r)].value
        a=wms['C{0}'.format(r)].value
        t=wms['D{0}'.format(r)].value
        root=Tk()
        root.geometry("1000x1000")
        print(q)
        image_path=r"C:\Users\Dell\Desktop\computer project bala 12d\QUESTION IMAGES\{0}.png".format(q)
        print(image_path)
        img=ImageTk.PhotoImage(Image.open(image_path))
        label=Label(root,image=img)
        label.pack(anchor=N)
        if t=='m':
            radio=StringVar()
            radio.set('A')
            values = [["A","A"],["B","B"],["C","C"],["D","D"]]
            iteration_no=0
            def clicked(option):
                 mathsmarkslstmcq.append([c,option,a,q])    
            for (text, val) in values:
                iteration_no+=1
                Radiobutton(root, text = text, variable = radio,value = val).pack(anchor=S)
            Button(root,text='save option',command=lambda: clicked(radio.get())).pack(anchor=S)
        elif t=='i':
            def submit():
                answer=entryvariable.get()
                mathsmarkslstint.append([c,answer,a,q])
            entryvariable=StringVar(root)
            answer_entry = Entry(root,textvariable = entryvariable)
            answer_entry.pack(anchor=S)
            enter_button=Button(root,text = 'ENTER', command = submit)
            enter_button.pack(anchor=S)
        button1=Button(root,text='next',command=root.destroy).pack(anchor=S)
        root.mainloop()
totalmarks=0
phymarks=0
chemmarks=0
mathmarks=0
root=Tk()
root.geometry("1600x1600")
label1=Label(root,text='QUESTION CODE',fg='brown', font=('Arial', 15))
label2=Label(root,text='CHAPTER',fg='brown', font=('Arial', 15))
label3=Label(root,text='YOUR ANSWER',fg='brown', font=('Arial', 15))
label4=Label(root,text='CORRECT ANSWER',fg='brown', font=('Arial', 15))
label5=Label(root,text='MARKS',fg='brown', font=('Arial', 15))
label1.grid(row=0,column=0)
label2.grid(row=0,column=3)
label3.grid(row=0,column=5)
label4.grid(row=0,column=7)
label5.grid(row=0,column=9)
cnt=0
for i in phymarkslstmcq:
     cnt+=1
     label6=Label(root,text=i[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=i[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=i[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=i[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if i[1]==i[2]:
         phymarks+=4
         label10=Label(root,text='+4',fg='green', font=('Arial', 15))
         label10.grid(row=cnt,column=9)
     elif i[1]=='':
         label11=Label(root,text='0',fg='orange', font=('Arial', 15))
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red', font=('Arial', 15))
         label12.grid(row=cnt,column=9)
         phymarks=phymarks-1
for j in phymarkslstint:
     cnt+=1
     label6=Label(root,text=j[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=j[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=j[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=j[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if j[1]==j[2]:
         phymarks+=4
         label10=Label(root,text='+4',fg='green')
         label10.grid(row=cnt,column=9)
     elif j[1]=='':
         label11=Label(root,text='0',fg='orange')
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red')
         label12.grid(row=cnt,column=9)
         phymarks=phymarks-1
for k in chemmarkslstmcq:
     cnt+=1
     label6=Label(root,text=k[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=k[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=k[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=k[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if k[1]==k[2]:
         chemmarks+=4
         label10=Label(root,text='+4',fg='green', font=('Arial', 15))
         label10.grid(row=cnt,column=9)
     elif k[1]=='':
         label11=Label(root,text='0',fg='orange', font=('Arial', 15))
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red', font=('Arial', 15))
         label12.grid(row=cnt,column=9)
         chemmarks=chemmarks-1
for l in chemmarkslstint:
     cnt+=1
     label6=Label(root,text=l[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=l[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=l[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=l[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if l[1]==l[2]:
         chemmarks+=4
         label10=Label(root,text='+4',fg='green', font=('Arial', 15))
         label10.grid(row=cnt,column=9)
     elif l[1]=='':
         label11=Label(root,text='0',fg='orange', font=('Arial', 15))
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red', font=('Arial', 15))
         label12.grid(row=cnt,column=9)
         chemmarks=chemmarks-1
for m in mathsmarkslstmcq:
     cnt+=1
     label6=Label(root,text=m[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=m[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=m[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=m[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if m[1]==m[2]:
         mathmarks+=4
         label10=Label(root,text='+4',fg='green', font=('Arial', 15))
         label10.grid(row=cnt,column=9)
     elif m[1]=='':
         label11=Label(root,text='0',fg='orange', font=('Arial', 15))
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red', font=('Arial', 15))
         label12.grid(row=cnt,column=9)
         mathmarks=mathmarks-1
for n in mathsmarkslstint:
     cnt+=1
     label6=Label(root,text=n[3],fg='black', font=('Arial', 15))
     label7=Label(root,text=n[0],fg='black', font=('Arial', 15))
     label8=Label(root,text=n[1],fg='black', font=('Arial', 15))
     label9=Label(root,text=n[2],fg='black', font=('Arial', 15))
     label6.grid(row=cnt,column=0)
     label7.grid(row=cnt,column=3)
     label8.grid(row=cnt,column=5)
     label9.grid(row=cnt,column=7)
     if n[1]==n[2]:
         mathmarks+=4
         label10=Label(root,text='+4',fg='green', font=('Arial', 15))
         label10.grid(row=cnt,column=9)
     elif n[1]=='':
         label11=Label(root,text='0',fg='orange', font=('Arial', 15))
         label11.grid(row=cnt,column=9)
     else:
         label12=Label(root,text='-1',fg='red', font=('Arial', 15))
         label12.grid(row=cnt,column=9)
         mathmarks=mathmarks-1

label6=Label(root,text='total physics:'+str(phymarks),fg='black', font=('Arial', 15))
label7=Label(root,text='total chemistry:'+str(chemmarks),fg='black', font=('Arial', 15))
label8=Label(root,text='total maths:'+str(mathmarks),fg='black', font=('Arial', 15))
label9=Label(root,text='total :'+str(phymarks+chemmarks+mathmarks),fg='black',font=('Arial', 15))
label6.grid(row=cnt+1,column=0)
label7.grid(row=cnt+1,column=3)
label8.grid(row=cnt+1,column=5)
label9.grid(row=cnt+1,column=7)
         
        
    
        

        
    
    
        
